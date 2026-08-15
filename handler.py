from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from states import user_state
from config import ADMIN_ID, SUPPORT_USERNAME, BOT_NAME, FORCE_JOIN_CHANNEL, FORCE_JOIN_LINK
from buttons import main_menu, shop_menu, deposit_menu, product_menu, quantity_menu, force_join_menu
from admin import admin_buttons
from database import create_user, get_balance, update_balance, add_order, get_orders, get_order_by_id, update_order_status, c, get_all_users, get_stock_count, take_codes, add_stock, add_referral, activate_referral_bonus, get_refer_stats, get_all_stock
from bot import bot
import io

def is_user_joined(bot, user_id):
    try:
        member = bot.get_chat_member(FORCE_JOIN_CHANNEL, user_id)
        return member.status in ['member', 'administrator', 'creator']
    except Exception as e:
        print(f"Join check error {user_id}: {e}")
        return False

def register_handlers(bot):

    @bot.message_handler(commands=["start"])
    def start(message):
        user_id = message.from_user.id
        args = message.text.split()
        if len(args) > 1:
            try:
                ref_id = int(args[1])
                if ref_id!= user_id:
                    create_user(user_id, ref_id)
                    add_referral(ref_id, user_id)
            except:
                create_user(user_id)
        else:
            create_user(user_id)

        if user_id!= ADMIN_ID and not is_user_joined(bot, user_id):
            bot.send_message(message.chat.id, f"⚠ Bot use korte hole amader channel e join korte hobe!\n\n📢 Channel: {FORCE_JOIN_CHANNEL}\n\nJoin kore Verify koro.", reply_markup=force_join_menu())
            return

        bot.send_message(message.chat.id, f"🤖 {BOT_NAME}\nWelcome to ProxyStore AI", reply_markup=main_menu())

    @bot.message_handler(commands=["admin"])
    def admin(message):
        if message.from_user.id!= ADMIN_ID:
            bot.reply_to(message, "❌ Access Denied")
            return
        bot.send_message(message.chat.id, "👑 Admin Panel", reply_markup=admin_buttons())

    @bot.callback_query_handler(func=lambda call: True)
    def callback(call):
        msg_id = call.message.message_id
        chat_id = call.message.chat.id
        user_id = call.from_user.id
        create_user(user_id)

        if call.data == "verify_join":
            if is_user_joined(bot, user_id):
                try: bot.edit_message_text(f"🤖 {BOT_NAME}\nWelcome to ProxyStore AI", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                except: pass
            else:
                bot.answer_callback_query(call.id, "❌ Tumi ekhono Channel e Join koro nai! Age Join koro.", show_alert=True)
            return

        if user_id!= ADMIN_ID and not is_user_joined(bot, user_id):
            try: bot.answer_callback_query(call.id, "⚠ Age Channel Join Koro!")
            except: pass
            bot.send_message(chat_id, f"⚠ Age Channel Join Koro!\n{FORCE_JOIN_CHANNEL}", reply_markup=force_join_menu())
            return

        if call.data == "shop":
            try: bot.edit_message_text("🛒 Select Category", chat_id=chat_id, message_id=msg_id, reply_markup=shop_menu())
            except: pass
        elif call.data == "vpn_list":
            try: bot.edit_message_text("🌐 VPN Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("vpn"))
            except: pass
        elif call.data == "proxy_list":
            try: bot.edit_message_text("🌍 Proxy Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("proxy"))
            except: pass
        elif call.data == "gmail_list":
            try: bot.edit_message_text("📧 Gmail Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("gmail"))
            except: pass
        elif call.data == "outlook_list":
            try: bot.edit_message_text("📮 Outlook Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("outlook"))
            except: pass
        elif call.data == "hotmail_list":
            try: bot.edit_message_text("📬 Hotmail Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("hotmail"))
        elif call.data == "edumail_list":
            try: bot.edit_message_text("📬 edumail Products", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("edumail"))
            except: pass
        elif call.data == "morelogin_list":
            try: bot.edit_message_text("🖥 Morelogin 100 Minutes", chat_id=chat_id, message_id=msg_id, reply_markup=product_menu("morelogin"))
            except: pass
        elif call.data == "noop":
            bot.answer_callback_query(call.id, "Quantity change korte + - use koro")
        elif call.data.startswith("select_qty|"):
            parts = call.data.split("|")
            category, name, price = parts[1], parts[2], float(parts[3])
            try: bot.edit_message_text(f"🛒 {name}\n💎 Price: {price} BDT\nStock: unlimited\nQuantity: 1", chat_id=chat_id, message_id=msg_id, reply_markup=quantity_menu(category, name, price, 1))
            except: pass
        elif call.data.startswith("qty_plus|"):
            parts = call.data.split("|")
            category, name, price, qty = parts[1], parts[2], float(parts[3]), int(parts[4])
            qty += 1
            try: bot.edit_message_text(f"🛒 {name}\n💎 Price: {price} BDT\nStock: unlimited\nQuantity: {qty}", chat_id=chat_id, message_id=msg_id, reply_markup=quantity_menu(category, name, price, qty))
            except: pass
        elif call.data.startswith("qty_minus|"):
            parts = call.data.split("|")
            category, name, price, qty = parts[1], parts[2], float(parts[3]), int(parts[4])
            if qty > 1: qty -= 1
            try: bot.edit_message_text(f"🛒 {name}\n💎 Price: {price} BDT\nStock: unlimited\nQuantity: {qty}", chat_id=chat_id, message_id=msg_id, reply_markup=quantity_menu(category, name, price, qty))
            except: pass
        elif call.data.startswith("custom_qty|"):
            parts = call.data.split("|")
            category, name, price = parts[1], parts[2], float(parts[3])
            user_state[user_id] = {"step": "custom_qty", "category": category, "name": name, "price": price}
            bot.send_message(chat_id, "📝 Koyta niba? Number likhe pathao")
        elif call.data.startswith("buy|"):
            parts = call.data.split("|")
            category, name, price, qty = parts[1], parts[2], float(parts[3]), int(parts[4])
            total_price = price * qty
            balance = get_balance(user_id)
            if balance >= total_price:
                is_auto = False
                if category == "morelogin": is_auto = True
                elif category == "proxy" and "owl" in name.lower(): is_auto = True
                if is_auto:
                    available = get_stock_count(category, name)
                    if available < qty:
                        bot.send_message(ADMIN_ID, f"⚠ Stock sesh! {name} - {qty} pcs order asche but stock {available} pcs")
                        try: bot.edit_message_text(f"❌ Stock e nai. Admin ke janao. Stock: {available} pcs", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                        except: pass
                        return
                update_balance(user_id, -total_price)
                if is_auto:
                    codes = take_codes(category, name, qty)
                    import openpyxl
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Delivery"
                    ws.append(["Product Name", "No", "Account Details"])
                    ws.column_dimensions['A'].width = 30
                    ws.column_dimensions['B'].width = 10
                    ws.column_dimensions['C'].width = 90
                    for i, code in enumerate(codes, 1): ws.append([name, i, code])
                    file_stream = io.BytesIO()
                    wb.save(file_stream)
                    file_stream.seek(0)
                    file_stream.name = f"{name.replace(' ','_')}_{qty}pcs.xlsx"
                    add_order(user_id, f"{name} x{qty}", total_price)
                    bot.send_document(user_id, file_stream, caption=f"✅ Order Delivered!\n\nProduct: {name}\nQuantity: {qty} pcs\nTotal: {total_price} BDT\n\nProblem hole {SUPPORT_USERNAME}")
                    try: bot.edit_message_text(f"✅ Order Complete! File upore diye disi", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                    except: pass
                else:
                    oid = add_order(user_id, f"{name} x{qty}", total_price)
                    update_order_status(oid, "Pending")
                    try: bot.edit_message_text(f"✅ Order Confirmed!\n\nProduct: {name}\nQuantity: {qty} pcs\nTotal: {total_price} BDT\n\nAdmin 5-10 min er moddhe code diye dibe", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
                    except: pass
                    bot.send_message(ADMIN_ID, f"🛒 New Manual Order\nOrder ID: {oid}\nUser: {user_id}\nProduct: {name} x{qty}\nTotal: {total_price} BDT\n\n/admin > Pending Orders")
            else:
                try: bot.edit_message_text(f"❌ Not Enough Balance\nYour Balance: {balance} BDT\nRequired: {total_price} BDT", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu())
                except: pass

        # --- DEPOSIT FIXED PART ---
        elif call.data == "deposit":
            try: bot.edit_message_text("💰 Deposit Balance\nSelect Payment Method", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu())
            except: pass
        elif call.data == "bkash":
            try: bot.edit_message_text("💳 bKash Personal\n`01603940061`\n\n1. Number copy kore payment korun\n2. Payment er por 'Submit Payment' e click korun", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="Markdown")
            except: pass
        elif call.data == "nagad":
            try: bot.edit_message_text("💳 Nagad Personal\n`01603940061`\n\n1. Number copy kore payment korun\n2. Payment er por 'Submit Payment' e click korun", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="Markdown")
            except: pass
        elif call.data == "rocket":
            try: bot.edit_message_text("💳 Rocket Personal\n`off ase akon`\n\n1. Number copy kore payment korun\n2. Payment er por 'Submit Payment' e click korun", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="Markdown")
            except: pass
        elif call.data == "usdt":
            try: bot.edit_message_text("💲 USDT Payment\n\n🔹 TRC20:\n`TGE8oPaj7cYP14xuoHTZT19KxwSf12FYoz`\n\n🔹 BEP20:\n`0x0Bc20843c4452C6fAcAf7E1b757a00c0F79D6268`\n\n1. Address copy kore payment korun\n2. Payment er por 'Submit Payment' e click korun", chat_id=chat_id, message_id=msg_id, reply_markup=deposit_menu(), parse_mode="Markdown")
            except: pass
        elif call.data == "submit_payment":
            user_state[user_id] = {"step": "amount"}
            bot.send_message(chat_id, "💰 Enter Deposit Amount")

        elif call.data == "admin_add_stock":
            if user_id!= ADMIN_ID: return
            bot.send_message(chat_id, "📦 File pathao\n.txt ba.xlsx")
            user_state[user_id] = {"step": "wait_txt_file"}
        elif call.data == "admin_broadcast":
            if user_id!= ADMIN_ID: return
            bot.send_message(chat_id, "📢 Broadcast message likhe pathao.")
            user_state[user_id] = {"step": "broadcast_msg"}
        elif call.data == "admin_add_balance":
            if user_id!= ADMIN_ID: return
            bot.send_message(chat_id, "👤 User er Telegram ID dao")
            user_state[user_id] = {"step": "admin_user_id"}
        elif call.data == "admin_orders":
            if user_id!= ADMIN_ID: return
            try:
                c.execute("SELECT id, user_id, product, price, status FROM orders ORDER BY id DESC LIMIT 20")
                orders = c.fetchall()
                text = "📦 No orders yet" if not orders else "📦 Last 20 Orders\n"+"\n".join([f"ID: {x[0]} | User: {x[1]}\nProduct: {x[2]}\nPrice: {x[3]} BDT | {x[4]}\n" for x in orders])
                bot.send_message(chat_id, text)
            except Exception as e:
                from database import conn; conn.rollback()
                bot.send_message(chat_id, f"Error: {e}")
        elif call.data == "admin_stock_list":
            if user_id!= ADMIN_ID: return
            stocks = get_all_stock()
            if not stocks:
                bot.send_message(chat_id, "📦 Stock khali")
                return
            text = "📊 **Stock List**\n\n"
            for cat, prod, count in stocks: text += f"• {cat} | {prod} : {count} pcs\n"
            bot.send_message(chat_id, text, parse_mode="Markdown")
        elif call.data == "admin_pending":
            if user_id!= ADMIN_ID: return
            try:
                c.execute("SELECT id, user_id, product, price FROM orders WHERE status='Pending' ORDER BY id DESC")
                orders = c.fetchall()
                if not orders:
                    bot.send_message(chat_id, "✅ No Pending Orders")
                    return
                for o in orders:
                    markup = InlineKeyboardMarkup()
                    markup.add(InlineKeyboardButton("✅ Approve & Send Code", callback_data=f"approve_{o[0]}"))
                    bot.send_message(chat_id, f"🛒 Order ID: {o[0]}\nUser: {o[1]}\nProduct: {o[2]}\nPrice: {o[3]} BDT", reply_markup=markup)
            except Exception as e:
                from database import conn; conn.rollback()
                bot.send_message(chat_id, f"Error: {e}")
        elif call.data.startswith("approve_"):
            if user_id!= ADMIN_ID: return
            order_id = int(call.data.split("_")[1])
            bot.send_message(chat_id, f"📦 Enter Product Code for Order {order_id}")
            user_state[user_id] = {"step": "admin_code", "order_id": order_id}
        elif call.data.startswith("confirm_"):
            if user_id!= ADMIN_ID: return
            parts = call.data.split("_")
            target_user = int(parts[1]); amount = float(parts[2])
            update_balance(target_user, amount)
            new_balance = get_balance(target_user)
            bonus_to = activate_referral_bonus(target_user, amount)
            if bonus_to:
                try: bot.send_message(bonus_to, f"🎉 Refer Bonus! {target_user} {amount:.0f} BDT deposit korse, tai tumi 0.50 BDT paiso!")
                except: pass
            bot.send_message(target_user, f"✅ ডিপোজিট সফল!\n\n💰 যোগ হয়েছে: +{amount:.2f} টাকা\n💳 ব্যালেন্স: {new_balance:.2f} টাকা")
            try: bot.edit_message_text(f"✅ Confirmed. {amount} BDT added to {target_user}", chat_id=chat_id, message_id=msg_id)
            except: pass
        elif call.data.startswith("cancel_"):
            if user_id!= ADMIN_ID: return
            target_user = int(call.data.split("_")[1])
            bot.send_message(target_user, "⚠ Payment Verification Failed")
            try: bot.edit_message_text("❌ Cancelled by Admin", chat_id=chat_id, message_id=msg_id)
            except: pass
        elif call.data == "wallet":
            try: bot.edit_message_text(f"👛 Wallet\n💰 Balance: {get_balance(user_id)} BDT", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
            except: pass
        elif call.data == "orders":
            orders = get_orders(user_id)
            text = "📦 My Orders\nNo orders yet" if not orders else "📦 My Orders\n"+"\n".join([f"• {x[0]} - {x[1]} BDT - {x[2]}" for x in orders])
            try: bot.edit_message_text(text, chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
            except: pass
        elif call.data == "support":
            support_text = "🛡 **SUPPORT UPDATE**\n\n📩 সমস্যা হলে Support: @PolasChandra\n\n🕒 24/7 Assistance"
            try: bot.edit_message_text(support_text, chat_id=chat_id, message_id=msg_id, reply_markup=main_menu(), parse_mode="Markdown")
            except: pass
        elif call.data == "about":
            try: bot.edit_message_text(f"ℹ About {BOT_NAME}", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
            except: pass
        elif call.data == "home":
            try: bot.edit_message_text(f"🤖 {BOT_NAME}\nWelcome to ProxyStore AI", chat_id=chat_id, message_id=msg_id, reply_markup=main_menu())
            except: pass
        elif call.data == "refer":
            count, earn = get_refer_stats(user_id)
            bot_username = bot.get_me().username
            link = f"https://t.me/{bot_username}?start={user_id}"
            text = f"👥 **Refer & Earn**\n\n🔗 Tomar Link:\n`{link}`\n\n👤 Total Refer: {count}\n💰 Earn: {earn:.2f} BDT"
            try: bot.edit_message_text(text, chat_id=chat_id, message_id=msg_id, reply_markup=main_menu(), parse_mode="Markdown")
            except: pass
        try: bot.answer_callback_query(call.id)
        except: pass

    @bot.message_handler(func=lambda m: m.from_user.id in user_state, content_types=['text', 'document'])
    def process_all(message):
        user_id = message.from_user.id
        if user_id!= ADMIN_ID and not is_user_joined(bot, user_id):
            bot.send_message(message.chat.id, f"⚠ Age Channel Join Koro!\n{FORCE_JOIN_CHANNEL}", reply_markup=force_join_menu())
            return
        state = user_state.get(user_id)
        if not state: return
        if state["step"] == "wait_txt_file":
            if message.content_type == 'document' and message.document:
                try:
                    file_name = message.document.file_name.lower()
                    codes = []
                    file_info = bot.get_file(message.document.file_id)
                    downloaded_file = bot.download_file(file_info.file_path)
                    if file_name.endswith(".txt"):
                        codes = downloaded_file.decode("utf-8", errors="ignore").splitlines()
                        codes = [c.strip() for c in codes if c.strip()!= ""]
                    elif file_name.endswith(".xlsx"):
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(downloaded_file))
                        ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            if row and row[0]: codes.append(str(row[0]).strip())
                    else: bot.send_message(message.chat.id, "❌ Sudhu.txt ba.xlsx"); return
                    if not codes: bot.send_message(message.chat.id, "❌ File faka!"); return
                    state["codes"] = codes
                    state["step"] = "stock_category"
                    bot.send_message(message.chat.id, f"✅ {len(codes)} ta code peyechi\n\nCategory: `proxy` / `morelogin`", parse_mode="Markdown")
                except Exception as e: bot.send_message(message.chat.id, f"❌ Error: {e}")
            else: bot.send_message(message.chat.id, "❌ File upload koro")
        elif state["step"] == "stock_category":
            cat = message.text.lower().strip()
            if cat not in ["proxy", "morelogin"]: bot.send_message(message.chat.id, "❌ `proxy` ba `morelogin` likho", parse_mode="Markdown"); return
            state["category"] = cat
            state["step"] = "stock_product"
            bot.send_message(message.chat.id, "Product name dao")
        elif state["step"] == "stock_product":
            add_stock(state["category"], message.text.strip(), state["codes"])
            bot.send_message(message.chat.id, f"✅ Stock Add Done!\n{state['category']} | {message.text.strip()} : {len(state['codes'])} pcs")
            del user_state[user_id]
        elif state["step"] == "admin_user_id":
            state["target_id"] = int(message.text); state["step"] = "admin_amount"; bot.send_message(message.chat.id, "💰 Koto BDT?")
        elif state["step"] == "admin_amount":
            amount = float(message.text); target_id = state["target_id"]; update_balance(target_id, amount)
            bot.send_message(message.chat.id, f"✅ {target_id} ke {amount} BDT add"); bot.send_message(target_id, f"🎉 Admin {amount} BDT add korse"); del user_state[user_id]
        elif state["step"] == "broadcast_msg":
            all_users = get_all_users(); sent = 0
            for uid in all_users:
                try: bot.send_message(uid, f"📢 **Notice from {BOT_NAME}**\n\n{message.text}", parse_mode="Markdown"); sent += 1
                except: pass
            bot.send_message(message.chat.id, f"✅ Broadcast Done! {sent} jon ke pathano hoise"); del user_state[user_id]
        elif state["step"] == "admin_code":
            order_id = state["order_id"]; code = message.text; order = get_order_by_id(order_id)
            if not order: bot.send_message(message.chat.id, f"❌ Order {order_id} nai"); del user_state[user_id]; return
            user_to_send = order[1]; prod = order[2]; update_order_status(order_id, "Approved")
            bot.send_message(user_to_send, f"✅ Your Order Approved!\n\n📦 Product: {prod}\n🔑 Code:\n`{code}`", parse_mode="Markdown")
            bot.send_message(message.chat.id, f"✅ Order {order_id} Approved & Code sent to {user_to_send}"); del user_state[user_id]
        elif state["step"] == "amount":
            state["amount"] = message.text; state["step"] = "trx"; bot.send_message(message.chat.id, "🧾 Send Transaction ID / TrxID")
        elif state["step"] == "trx":
            amount = state['amount']; trx = message.text
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton("✅ Confirm", callback_data=f"confirm_{user_id}_{amount}"), InlineKeyboardButton("❌ Cancel", callback_data=f"cancel_{user_id}"))
            bot.send_message(ADMIN_ID,f"💰 New Deposit Request\n👤 {message.from_user.first_name}\n🆔 {user_id}\nAmount: {amount} BDT\nTRX ID: {trx}", reply_markup=markup)
            bot.send_message(message.chat.id,"✅ Deposit Request Sent."); del user_state[user_id]
        elif state["step"] == "custom_qty":
            try:
                qty = int(message.text)
                if qty < 1: qty = 1
                bot.send_message(message.chat.id, f"🛒 {state['name']}\nQuantity: {qty}", reply_markup=quantity_menu(state['category'], state['name'], state['price'], qty))
                del user_state[user_id]
            except: bot.send_message(message.chat.id, "❌ Sothik number dao")
